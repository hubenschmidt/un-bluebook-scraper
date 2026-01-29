#!/usr/bin/env python3
"""Migrate UN Blue Book CSV and Fellowship XLSX to SQLite database."""

import sqlite3
import csv
import os

DB_FILE = "un_data.db"
FELLOWSHIP_FILE = "UN_Disarmament_Fellowship_1979.xlsx"


def create_database():
    """Create SQLite database and tables."""
    conn = sqlite3.connect(DB_FILE)
    cur = conn.cursor()

    cur.execute("DROP TABLE IF EXISTS bluebook")
    cur.execute("DROP TABLE IF EXISTS fellowship")

    cur.execute("""
        CREATE TABLE bluebook (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            country TEXT,
            last_name TEXT,
            first_name TEXT,
            title TEXT,
            rank TEXT,
            rank_display TEXT,
            function TEXT,
            status TEXT,
            appointment TEXT,
            cred_presented TEXT,
            email TEXT,
            phone TEXT,
            mission TEXT,
            position INTEGER
        )
    """)

    cur.execute("""
        CREATE TABLE fellowship (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER,
            region TEXT,
            country TEXT,
            salutation TEXT,
            last_name TEXT,
            first_name TEXT,
            note TEXT
        )
    """)

    conn.commit()
    return conn


def import_bluebook(conn):
    """Import Blue Book CSV data."""
    cur = conn.cursor()

    with open("un_bluebook_representatives.csv", "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = []
        for row in reader:
            rows.append((
                row.get("_country", ""),
                row.get("BB_LastName", ""),
                row.get("BB_FirstName", ""),
                row.get("BB_Title", ""),
                row.get("BB_Dipl_Rank", ""),
                row.get("BB_Dipl_Rank_Display", ""),
                row.get("BB_Function", ""),
                row.get("BB_Status", ""),
                row.get("BB_Appointment", ""),
                row.get("BB_Cred_Presented", ""),
                row.get("BB_Email", ""),
                row.get("BB_PhnNumber", ""),
                row.get("BB_Mission", ""),
                row.get("BB_Position", "") or None
            ))

        cur.executemany("""
            INSERT INTO bluebook (country, last_name, first_name, title, rank, rank_display,
                                  function, status, appointment, cred_presented, email, phone,
                                  mission, position)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)

    conn.commit()
    return len(rows)


def import_fellowship(conn):
    """Import Fellowship XLSX data."""
    if not os.path.exists(FELLOWSHIP_FILE):
        print(f"  Skipping: {FELLOWSHIP_FILE} not found")
        return 0

    import openpyxl
    cur = conn.cursor()

    wb = openpyxl.load_workbook(FELLOWSHIP_FILE)
    sheet = wb.active

    rows = []
    for row_num in range(2, sheet.max_row + 1):  # Skip header
        year = sheet.cell(row_num, 1).value
        region = sheet.cell(row_num, 2).value
        country = sheet.cell(row_num, 3).value
        salutation = sheet.cell(row_num, 4).value
        last_name = sheet.cell(row_num, 5).value
        first_name = sheet.cell(row_num, 6).value
        note = sheet.cell(row_num, 7).value

        if country and last_name and str(year) != "Year":  # Skip header if present
            rows.append((
                int(float(year)) if year and str(year).replace('.', '').isdigit() else None,
                str(region).strip() if region else "",
                str(country).strip() if country else "",
                str(salutation).strip() if salutation else "",
                str(last_name).strip() if last_name else "",
                str(first_name).strip() if first_name else "",
                str(note).strip() if note else ""
            ))

    cur.executemany("""
        INSERT INTO fellowship (year, region, country, salutation, last_name, first_name, note)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, rows)

    conn.commit()
    return len(rows)


def find_matches(conn):
    """Find matches between fellowship alumni and current Blue Book representatives."""
    cur = conn.cursor()

    query = """
        SELECT
            f.year AS fellowship_year,
            f.country AS fellowship_country,
            f.salutation AS fellowship_salutation,
            f.last_name AS fellowship_last_name,
            f.first_name AS fellowship_first_name,
            b.country AS bluebook_country,
            b.title AS bluebook_title,
            b.first_name AS bluebook_first_name,
            b.last_name AS bluebook_last_name,
            b.rank AS bluebook_rank,
            b.function AS bluebook_function,
            b.status AS bluebook_status
        FROM fellowship f
        JOIN bluebook b ON
            UPPER(TRIM(f.country)) = UPPER(TRIM(b.country))
            AND UPPER(TRIM(f.last_name)) = UPPER(TRIM(b.last_name))
            AND UPPER(SUBSTR(TRIM(f.first_name), 1, 4)) = UPPER(SUBSTR(TRIM(b.first_name), 1, 4))
        WHERE b.status = 'Active'
        ORDER BY f.country, f.last_name
    """

    cur.execute(query)
    return cur.fetchall(), [desc[0] for desc in cur.description]


if __name__ == "__main__":
    print("Creating SQLite database...")
    conn = create_database()

    print("Importing Blue Book CSV...")
    bluebook_count = import_bluebook(conn)
    print(f"  Imported {bluebook_count} rows into 'bluebook' table")

    print("Importing Fellowship XLSX...")
    fellowship_count = import_fellowship(conn)
    if fellowship_count > 0:
        print(f"  Imported {fellowship_count} rows into 'fellowship' table")

        print("\nFinding matches...")
        matches, columns = find_matches(conn)
        print(f"  Found {len(matches)} matches\n")

        print("=" * 100)
        print("MATCHES: Fellowship Alumni Currently in UN Blue Book")
        print("=" * 100)

        for match in matches:
            print(f"\n{match[3]}, {match[4]} ({match[0]}, {match[1]})")
            print(f"  Now: {match[6]} {match[7]} {match[8]} - {match[10]} ({match[9]})")

    conn.close()
    print(f"\nDatabase saved to: {DB_FILE}")
