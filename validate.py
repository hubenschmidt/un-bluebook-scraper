#!/usr/bin/env python3
"""Validate matching results by comparing SQL and Python approaches."""

import sqlite3
import csv
import os

FELLOWSHIP_FILE = "UN_Disarmament_Fellowship_1979.xlsx"


def sql_matches():
    """Get matches using SQLite."""
    conn = sqlite3.connect("un_data.db")
    cur = conn.cursor()

    cur.execute("""
        SELECT
            f.year, f.country, f.last_name, f.first_name,
            b.country, b.last_name, b.first_name, b.function, b.rank
        FROM fellowship f
        JOIN bluebook b ON
            UPPER(TRIM(f.country)) = UPPER(TRIM(b.country))
            AND UPPER(TRIM(f.last_name)) = UPPER(TRIM(b.last_name))
            AND UPPER(SUBSTR(TRIM(f.first_name), 1, 4)) = UPPER(SUBSTR(TRIM(b.first_name), 1, 4))
        WHERE b.status = 'Active'
        ORDER BY f.country, f.last_name
    """)

    results = set()
    for row in cur.fetchall():
        key = (row[1].upper(), row[2].upper(), row[3][:4].upper().strip())
        results.add(key)

    conn.close()
    return results


def python_matches():
    """Get matches using pure Python."""
    if not os.path.exists(FELLOWSHIP_FILE):
        return None

    import openpyxl

    # Load Blue Book
    bluebook = []
    with open("un_bluebook_representatives.csv", "r", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if row.get("BB_Status") == "Active" and row.get("_country") and row.get("BB_LastName"):
                bluebook.append({
                    "country": row["_country"].strip(),
                    "last_name": row["BB_LastName"].strip(),
                    "first_name": row.get("BB_FirstName", "").strip()
                })

    # Load Fellowship
    fellowship = []
    wb = openpyxl.load_workbook(FELLOWSHIP_FILE)
    sheet = wb.active
    for row_num in range(2, sheet.max_row + 1):
        country = sheet.cell(row_num, 3).value
        last_name = sheet.cell(row_num, 5).value
        first_name = sheet.cell(row_num, 6).value
        if country and last_name and str(sheet.cell(row_num, 1).value) != "Year":
            fellowship.append({
                "country": str(country).strip(),
                "last_name": str(last_name).strip(),
                "first_name": str(first_name).strip() if first_name else ""
            })

    # Build lookup
    bb_lookup = {}
    for bp in bluebook:
        key = (bp["country"].upper(), bp["last_name"].upper())
        bb_lookup.setdefault(key, []).append(bp)

    # Find matches
    results = set()
    for fp in fellowship:
        key = (fp["country"].upper(), fp["last_name"].upper())
        for bp in bb_lookup.get(key, []):
            fp_first = fp["first_name"][:4].upper().strip()
            bp_first = bp["first_name"][:4].upper().strip()
            if fp_first and bp_first and fp_first == bp_first:
                results.add((fp["country"].upper(), fp["last_name"].upper(), fp_first))

    return results


if __name__ == "__main__":
    if not os.path.exists(FELLOWSHIP_FILE):
        print(f"Skipping validation: {FELLOWSHIP_FILE} not found")
        print("To run validation, place the fellowship XLSX file in this directory.")
        exit(0)

    print("Running SQL matching...")
    sql_results = sql_matches()
    print(f"  SQL found: {len(sql_results)} matches")

    print("\nRunning Python matching...")
    py_results = python_matches()
    print(f"  Python found: {len(py_results)} matches")

    print("\n" + "=" * 50)
    if sql_results == py_results:
        print("✓ VALIDATED: Both methods found identical matches")
        print(f"  Total matches: {len(sql_results)}")
    else:
        print("✗ MISMATCH DETECTED")
        only_sql = sql_results - py_results
        only_py = py_results - sql_results
        if only_sql:
            print(f"  Only in SQL: {only_sql}")
        if only_py:
            print(f"  Only in Python: {only_py}")
