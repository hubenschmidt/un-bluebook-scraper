#!/usr/bin/env python3
"""Python-based matching between Fellowship and Blue Book data (for validation)."""

import csv
import os

FELLOWSHIP_FILE = "UN_Disarmament_Fellowship_1979.xlsx"


def load_bluebook():
    """Load Blue Book CSV data."""
    people = []
    with open("un_bluebook_representatives.csv", "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            country = row.get("_country", "").strip()
            last_name = row.get("BB_LastName", "").strip()
            first_name = row.get("BB_FirstName", "").strip()
            status = row.get("BB_Status", "").strip()

            if country and last_name and status == "Active":
                people.append({
                    "country": country,
                    "last_name": last_name,
                    "first_name": first_name,
                    "title": row.get("BB_Title", ""),
                    "rank": row.get("BB_Dipl_Rank", ""),
                    "function": row.get("BB_Function", ""),
                    "status": status
                })
    return people


def load_fellowship():
    """Load Fellowship XLSX data."""
    if not os.path.exists(FELLOWSHIP_FILE):
        return None

    import openpyxl
    people = []
    wb = openpyxl.load_workbook(FELLOWSHIP_FILE)
    sheet = wb.active

    for row_num in range(2, sheet.max_row + 1):
        year = sheet.cell(row_num, 1).value
        region = sheet.cell(row_num, 2).value
        country = sheet.cell(row_num, 3).value
        salutation = sheet.cell(row_num, 4).value
        last_name = sheet.cell(row_num, 5).value
        first_name = sheet.cell(row_num, 6).value

        if country and last_name and str(year) != "Year":
            people.append({
                "year": int(float(year)) if year else None,
                "region": str(region).strip() if region else "",
                "country": str(country).strip(),
                "salutation": str(salutation).strip() if salutation else "",
                "last_name": str(last_name).strip(),
                "first_name": str(first_name).strip() if first_name else ""
            })
    return people


def find_matches(fellowship, bluebook):
    """Find matches using Python (same logic as SQL)."""
    # Build lookup by (country_upper, last_name_upper)
    bb_lookup = {}
    for bp in bluebook:
        key = (bp["country"].upper(), bp["last_name"].upper())
        if key not in bb_lookup:
            bb_lookup[key] = []
        bb_lookup[key].append(bp)

    matches = []
    for fp in fellowship:
        key = (fp["country"].upper(), fp["last_name"].upper())
        candidates = bb_lookup.get(key, [])

        for bp in candidates:
            # Match first 4 chars of first name
            fp_first = fp["first_name"][:4].upper().strip()
            bp_first = bp["first_name"][:4].upper().strip()

            if fp_first and bp_first and fp_first == bp_first:
                matches.append({
                    "fellowship_year": fp["year"],
                    "fellowship_country": fp["country"],
                    "fellowship_salutation": fp["salutation"],
                    "fellowship_last_name": fp["last_name"],
                    "fellowship_first_name": fp["first_name"],
                    "bluebook_country": bp["country"],
                    "bluebook_title": bp["title"],
                    "bluebook_first_name": bp["first_name"],
                    "bluebook_last_name": bp["last_name"],
                    "bluebook_rank": bp["rank"],
                    "bluebook_function": bp["function"],
                    "bluebook_status": bp["status"]
                })
    return matches


if __name__ == "__main__":
    print("Loading Blue Book data...")
    bluebook = load_bluebook()
    print(f"  Loaded {len(bluebook)} active representatives")

    print("Loading Fellowship data...")
    fellowship = load_fellowship()

    if fellowship is None:
        print(f"  Skipping: {FELLOWSHIP_FILE} not found")
        print("\nTo run matching, place the fellowship XLSX file in this directory.")
        exit(0)

    print(f"  Loaded {len(fellowship)} fellowship alumni")

    print("\nFinding matches (Python)...")
    matches = find_matches(fellowship, bluebook)
    print(f"  Found {len(matches)} matches\n")

    print("=" * 100)
    print("MATCHES: Fellowship Alumni Currently in UN Blue Book (Python)")
    print("=" * 100)

    for m in sorted(matches, key=lambda x: (x["fellowship_country"], x["fellowship_last_name"])):
        print(f"\n{m['fellowship_last_name']}, {m['fellowship_first_name']} ({m['fellowship_year']}, {m['fellowship_country']})")
        print(f"  Now: {m['bluebook_title']} {m['bluebook_first_name']} {m['bluebook_last_name']} - {m['bluebook_function']} ({m['bluebook_rank']})")

    # Save to CSV
    if matches:
        with open("matches_python.csv", "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=matches[0].keys())
            writer.writeheader()
            writer.writerows(matches)
        print(f"\nSaved to: matches_python.csv")
